"""
Business Plan (BP) router — CAS team
"""
from typing import Optional, Annotated
from datetime import date, datetime
import io
import base64

from fastapi import APIRouter, HTTPException, UploadFile, File, Depends
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.deps import DB, CurrentUser, LeaderOrAdmin, get_current_user
from app.models.user import User
from app.models.business_plan import (
    BusinessPlan, BPLine, BPActivity, BPExcelAnalysis, BPRecommendation,
    BPLineCategory, BPActivityStatus,
    BPChecklist, BPComment, BPMilestone,
)
from app.models.business import Business
from app.models.business_intel import PremisaNegocio
from app.models.notification import Notification
from app.schemas.business_plan import (
    BusinessPlanCreate, BusinessPlanUpdate,
    BPLineCreate, BPLineUpdate,
    BPActivityCreate, BPActivityUpdate,
    BPRecommendationCreate, BPRecommendationUpdate,
    BPChecklistItemCreate, BPChecklistItemUpdate,
    BPCommentCreate,
    BPMilestoneCreate, BPMilestoneUpdate,
)

router = APIRouter(prefix="/bp", tags=["Business Plan"])

# Supported image MIME types
IMAGE_MIME_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".gif": "image/gif",
    ".webp": "image/webp",
}

EXCEL_EXTENSIONS = (".xlsx", ".xls", ".xlsm")


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _compute_annual_plan(monthly_plan: Optional[dict]) -> Optional[float]:
    """Sum all monthly values to get annual plan total."""
    if not monthly_plan:
        return None
    try:
        return sum(float(v) for v in monthly_plan.values() if v is not None)
    except (TypeError, ValueError):
        return None


def _is_overdue(activity: BPActivity) -> bool:
    if activity.status in (BPActivityStatus.COMPLETADA, BPActivityStatus.CANCELADA):
        return False
    if activity.due_date and activity.due_date < date.today():
        return True
    return False


def _bp_to_dict(bp: BusinessPlan, activities_stats: Optional[dict] = None) -> dict:
    stats = activities_stats or {}
    return {
        "id": bp.id,
        "business_id": bp.business_id,
        "business_name": bp.business.name if bp.business else None,
        "year": bp.year,
        "status": bp.status.value if hasattr(bp.status, "value") else bp.status,
        "version": bp.version,
        "name": bp.name,
        "description": bp.description,
        "scope": bp.scope,
        "total_ingresos_plan": bp.total_ingresos_plan,
        "total_costos_plan": bp.total_costos_plan,
        "margen_bruto_plan": bp.margen_bruto_plan,
        "created_by_id": bp.created_by_id,
        "created_by_name": bp.created_by.full_name if bp.created_by else None,
        "created_at": bp.created_at,
        "updated_at": bp.updated_at,
        "activities_total": stats.get("total", 0),
        "activities_completed": stats.get("completed", 0),
        "activities_overdue": stats.get("overdue", 0),
    }


def _line_to_dict(line: BPLine) -> dict:
    return {
        "id": line.id,
        "bp_id": line.bp_id,
        "business_id": line.business_id,
        "business_name": line.business.name if hasattr(line, 'business') and line.business else None,
        "premisa_id": line.premisa_id,
        "premisa_title": (line.premisa.title if hasattr(line, 'premisa') and line.premisa else None),
        "category": line.category.value if hasattr(line.category, "value") else line.category,
        "subcategory": line.subcategory,
        "name": line.name,
        "unit": line.unit,
        "monthly_plan": line.monthly_plan,
        "monthly_actual": line.monthly_actual,
        "annual_plan": line.annual_plan,
        "annual_actual": line.annual_actual,
        "notes": line.notes,
        "order_index": line.order_index,
        "is_ai_generated": line.is_ai_generated,
        "ai_confidence": line.ai_confidence,
        "ai_rationale": line.ai_rationale,
    }


def _activity_to_dict(act: BPActivity) -> dict:
    overdue = _is_overdue(act)
    displayed_status = act.status.value if hasattr(act.status, "value") else act.status
    if overdue and act.status == BPActivityStatus.PENDIENTE:
        displayed_status = BPActivityStatus.VENCIDA.value
    checklist = act.checklist if hasattr(act, 'checklist') and act.checklist else []
    comments = act.comments if hasattr(act, 'comments') and act.comments else []
    return {
        "id": act.id,
        "bp_id": act.bp_id,
        "title": act.title,
        "description": act.description,
        "category": act.category.value if hasattr(act.category, "value") else act.category,
        "priority": act.priority.value if hasattr(act.priority, "value") else act.priority,
        "status": displayed_status,
        "owner_id": act.owner_id,
        "owner_name": act.owner.full_name if act.owner else None,
        "due_date": str(act.due_date) if act.due_date else None,
        "completion_date": str(act.completion_date) if act.completion_date else None,
        "progress": act.progress,
        "premisa_id": act.premisa_id,
        "notes": act.notes,
        "evidence": act.evidence,
        "order_index": act.order_index,
        "is_overdue": overdue,
        "created_at": act.created_at,
        "updated_at": act.updated_at,
        # Schedule & Tasks fields
        "start_date": str(act.start_date) if act.start_date else None,
        "estimated_hours": act.estimated_hours,
        "actual_hours": act.actual_hours,
        "depends_on_id": act.depends_on_id,
        "is_milestone": act.is_milestone,
        "reminder_days_before": act.reminder_days_before,
        "tags": act.tags,
        "grupo": act.grupo,
        "checklist_total": len(checklist),
        "checklist_done": sum(1 for i in checklist if i.is_completed),
        "comment_count": len([c for c in comments if not c.is_deleted]),
    }


def _analysis_to_dict(a: BPExcelAnalysis) -> dict:
    return {
        "id": a.id,
        "bp_id": a.bp_id,
        "filename": a.filename,
        "file_size": a.file_size,
        "file_type": a.file_type,
        "parsed_data": a.parsed_data,
        "ai_summary": a.ai_summary,
        "ai_insights": a.ai_insights,
        "structured_extraction": a.structured_extraction,
        "applied_at": a.applied_at,
        "uploaded_by_id": a.uploaded_by_id,
        "uploaded_by_name": a.uploaded_by.full_name if a.uploaded_by else None,
        "uploaded_at": a.uploaded_at,
    }


def _recommendation_to_dict(r: BPRecommendation) -> dict:
    return {
        "id": r.id,
        "bp_id": r.bp_id,
        "source": r.source,
        "category": r.category,
        "title": r.title,
        "description": r.description,
        "priority": r.priority,
        "status": r.status,
        "impact_level": r.impact_level,
        "is_ai_generated": r.is_ai_generated,
        "rec_metadata": r.rec_metadata,
        "created_at": r.created_at,
        "updated_at": r.updated_at,
    }


async def _get_activity_stats(db: AsyncSession, bp_id: int) -> dict:
    result = await db.execute(
        select(BPActivity).where(
            BPActivity.bp_id == bp_id,
            BPActivity.is_deleted == False,
        )
    )
    activities = result.scalars().all()
    total = len(activities)
    completed = sum(1 for a in activities if a.status == BPActivityStatus.COMPLETADA)
    overdue = sum(1 for a in activities if _is_overdue(a))
    return {"total": total, "completed": completed, "overdue": overdue}


async def _call_gemini_structured(db: AsyncSession, bp_year: int, contents_parts: list) -> dict:
    """Call Gemini API with the structured BP extraction prompt. Returns parsed JSON dict."""
    target_year = bp_year + 1

    prompt = (
        f"Eres un analista experto en planes de negocio para Vanti (empresa de gas/energía en Colombia), equipo CAS.\n"
        f"Analiza el contenido del archivo del Plan de Negocio anterior y genera el nuevo BP proyectado para {target_year}.\n\n"
        "IMPORTANTE: Responde ÚNICAMENTE con JSON válido, sin texto adicional, sin markdown, sin bloques de código.\n\n"
        "Estructura requerida:\n"
        "{\n"
        f'  "year_suggested": {target_year},\n'
        '  "summary": "Resumen ejecutivo del análisis (3-4 oraciones con cifras concretas)",\n'
        '  "financial_lines": [\n'
        '    {\n'
        '      "category": "ingreso|costo_fijo|costo_variable|magnitud|margen",\n'
        '      "subcategory": "subcategoría (ej: Ventas, Nómina, Marketing)",\n'
        '      "name": "nombre descriptivo de la línea",\n'
        '      "unit": "COP|USD|%|clientes|unidades|contratos|usuarios",\n'
        '      "monthly_plan": {"1": valor_num, "2": valor_num, "3": valor_num, "4": valor_num, "5": valor_num, "6": valor_num, "7": valor_num, "8": valor_num, "9": valor_num, "10": valor_num, "11": valor_num, "12": valor_num},\n'
        '      "ai_confidence": 85,\n'
        '      "ai_rationale": "Proyectado con base en tendencia histórica +8%"\n'
        '    }\n'
        '  ],\n'
        '  "activities": [\n'
        '    {\n'
        '      "title": "Título accionable y específico",\n'
        '      "description": "Descripción breve de qué se debe hacer",\n'
        '      "category": "comercial|operativo|financiero|estrategico|regulatorio|tecnologia",\n'
        '      "priority": "critica|alta|media|baja",\n'
        '      "due_date": "YYYY-MM-DD o null",\n'
        '      "notes": "Contexto o dependencias"\n'
        '    }\n'
        '  ],\n'
        '  "recommendations": [\n'
        '    {\n'
        '      "category": "comercial|financiero|operativo|estrategico|riesgo|oportunidad",\n'
        '      "title": "Título conciso de la recomendación",\n'
        '      "description": "Descripción detallada con argumentos cuantitativos cuando sea posible",\n'
        '      "priority": "critica|alta|media|baja",\n'
        '      "impact_level": "alto|medio|bajo"\n'
        '    }\n'
        '  ],\n'
        '  "risks": ["descripción riesgo 1", "descripción riesgo 2"],\n'
        '  "opportunities": ["oportunidad 1", "oportunidad 2"]\n'
        '}'
    )

    try:
        from app.core.config import get_service_config_value
        import httpx

        api_key = await get_service_config_value(db, "gemini", "api_key")
        model_name = await get_service_config_value(db, "gemini", "model") or "gemini-1.5-flash"

        if not api_key:
            return _fallback_extraction("Gemini no configurado. Configure la API key en Configuración > Integraciones.")

        # Build the request parts: text prompt + any file content
        request_parts = contents_parts + [{"text": prompt}]

        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={api_key}",
                json={
                    "contents": [{"parts": request_parts}],
                    "generationConfig": {
                        "temperature": 0.3,
                        "maxOutputTokens": 8192,
                        "responseMimeType": "application/json",
                    },
                },
            )
            if resp.status_code != 200:
                return _fallback_extraction(f"Error Gemini {resp.status_code}: {resp.text[:200]}")

            rdata = resp.json()
            raw_text = (
                rdata.get("candidates", [{}])[0]
                .get("content", {})
                .get("parts", [{}])[0]
                .get("text", "")
            )

            # Try to parse JSON — strip any accidental markdown fences
            import json
            clean = raw_text.strip()
            if clean.startswith("```"):
                clean = clean.split("```")[1]
                if clean.startswith("json"):
                    clean = clean[4:]
                clean = clean.strip()
            if clean.endswith("```"):
                clean = clean[:-3].strip()

            parsed = json.loads(clean)
            return parsed

    except Exception as exc:
        return _fallback_extraction(f"Error al procesar con IA: {str(exc)[:200]}")


def _fallback_extraction(message: str) -> dict:
    return {
        "year_suggested": None,
        "summary": message,
        "financial_lines": [],
        "activities": [],
        "recommendations": [],
        "risks": [],
        "opportunities": [],
    }


# ─── Dashboard ────────────────────────────────────────────────────────────────

@router.get("/dashboard")
async def bp_dashboard(db: DB, user: CurrentUser, year: Optional[int] = None):
    """Summary stats: all businesses, completion %, overdue activities."""
    biz_result = await db.execute(select(Business))
    businesses = biz_result.scalars().all()

    bp_query = (
        select(BusinessPlan)
        .where(BusinessPlan.is_deleted == False)
        .options(selectinload(BusinessPlan.business), selectinload(BusinessPlan.created_by))
        .order_by(BusinessPlan.year.desc())
    )
    if year:
        bp_query = bp_query.where(BusinessPlan.year == year)
    bp_result = await db.execute(bp_query)
    all_bps = bp_result.scalars().all()

    # Latest BP per business
    bp_by_business: dict[int, BusinessPlan] = {}
    for bp in all_bps:
        if bp.business_id not in bp_by_business:
            bp_by_business[bp.business_id] = bp

    status_counts: dict[str, int] = {}
    for bp in all_bps:
        s = bp.status.value if hasattr(bp.status, "value") else str(bp.status)
        status_counts[s] = status_counts.get(s, 0) + 1

    # Fetch activities for all BPs
    bp_ids = [bp.id for bp in bp_by_business.values()]
    acts_by_bp: dict[int, list] = {}
    if bp_ids:
        act_result = await db.execute(
            select(BPActivity).where(
                BPActivity.bp_id.in_(bp_ids),
                BPActivity.is_deleted == False,
            )
        )
        for a in act_result.scalars().all():
            acts_by_bp.setdefault(a.bp_id, []).append(a)

    summaries = []
    for biz in businesses:
        bp = bp_by_business.get(biz.id)
        if bp:
            acts = acts_by_bp.get(bp.id, [])
            total_acts = len(acts)
            completed = sum(1 for a in acts if a.status == BPActivityStatus.COMPLETADA)
            overdue = sum(1 for a in acts if _is_overdue(a))
            completion_pct = round((completed / total_acts * 100) if total_acts > 0 else 0.0, 1)
            summaries.append({
                "business_id": biz.id,
                "business_name": biz.name,
                "business_color": getattr(biz, "color", None),
                "latest_bp_id": bp.id,
                "year": bp.year,
                "status": bp.status.value if hasattr(bp.status, "value") else bp.status,
                "total_ingresos_plan": bp.total_ingresos_plan,
                "total_costos_plan": bp.total_costos_plan,
                "margen_bruto_plan": bp.margen_bruto_plan,
                "activities_total": total_acts,
                "activities_completed": completed,
                "activities_overdue": overdue,
                "completion_pct": completion_pct,
            })
        else:
            summaries.append({
                "business_id": biz.id,
                "business_name": biz.name,
                "business_color": getattr(biz, "color", None),
                "latest_bp_id": None,
                "year": None,
                "status": None,
                "total_ingresos_plan": None,
                "total_costos_plan": None,
                "margen_bruto_plan": None,
                "activities_total": 0,
                "activities_completed": 0,
                "activities_overdue": 0,
                "completion_pct": 0.0,
            })

    all_acts_flat = [a for acts in acts_by_bp.values() for a in acts]
    return {
        "total_bps": len(all_bps),
        "total_businesses_with_bp": len(bp_by_business),
        "total_activities": len(all_acts_flat),
        "total_overdue": sum(1 for a in all_acts_flat if _is_overdue(a)),
        "by_status": status_counts,
        "businesses": summaries,
    }


# ─── BusinessPlan CRUD ────────────────────────────────────────────────────────

@router.get("")
async def list_bps(
    db: DB, user: CurrentUser,
    year: Optional[int] = None,
    business_id: Optional[int] = None,
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
):
    query = (
        select(BusinessPlan)
        .where(BusinessPlan.is_deleted == False)
        .options(selectinload(BusinessPlan.business), selectinload(BusinessPlan.created_by))
    )
    if year:
        query = query.where(BusinessPlan.year == year)
    if business_id:
        query = query.where(BusinessPlan.business_id == business_id)
    if status:
        query = query.where(BusinessPlan.status == status)
    query = query.order_by(BusinessPlan.year.desc(), BusinessPlan.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(query)
    bps = result.scalars().all()
    out = []
    for bp in bps:
        stats = await _get_activity_stats(db, bp.id)
        out.append(_bp_to_dict(bp, stats))
    return out


@router.post("", status_code=201)
async def create_bp(payload: BusinessPlanCreate, db: DB, user: LeaderOrAdmin):
    existing = await db.execute(
        select(BusinessPlan).where(
            BusinessPlan.business_id == payload.business_id,
            BusinessPlan.year == payload.year,
            BusinessPlan.is_deleted == False,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Ya existe un BP para este negocio en {payload.year}")
    bp = BusinessPlan(created_by_id=user.id, **payload.model_dump())
    db.add(bp)
    await db.flush()
    result = await db.execute(
        select(BusinessPlan)
        .where(BusinessPlan.id == bp.id)
        .options(selectinload(BusinessPlan.business), selectinload(BusinessPlan.created_by))
    )
    bp = result.scalar_one()
    return _bp_to_dict(bp)


@router.get("/{bp_id}")
async def get_bp(bp_id: int, db: DB, user: CurrentUser):
    result = await db.execute(
        select(BusinessPlan)
        .where(BusinessPlan.id == bp_id, BusinessPlan.is_deleted == False)
        .options(
            selectinload(BusinessPlan.business),
            selectinload(BusinessPlan.created_by),
            selectinload(BusinessPlan.lines).selectinload(BPLine.business),
            selectinload(BusinessPlan.lines).selectinload(BPLine.premisa),
            selectinload(BusinessPlan.activities).selectinload(BPActivity.owner),
            selectinload(BusinessPlan.activities).selectinload(BPActivity.checklist),
            selectinload(BusinessPlan.activities).selectinload(BPActivity.comments),
            selectinload(BusinessPlan.excel_analyses).selectinload(BPExcelAnalysis.uploaded_by),
            selectinload(BusinessPlan.recommendations),
        )
    )
    bp = result.scalar_one_or_none()
    if not bp:
        raise HTTPException(status_code=404, detail="Business Plan no encontrado")

    data = _bp_to_dict(bp)
    data["lines"] = [_line_to_dict(l) for l in bp.lines if not l.is_deleted]
    data["activities"] = [_activity_to_dict(a) for a in bp.activities if not a.is_deleted]
    data["excel_analyses"] = [_analysis_to_dict(a) for a in bp.excel_analyses]
    data["recommendations"] = [_recommendation_to_dict(r) for r in bp.recommendations if not r.is_deleted]

    # Recompute totals from lines
    ingreso_lines = [l for l in bp.lines if not l.is_deleted and l.category == BPLineCategory.INGRESO]
    costo_lines = [l for l in bp.lines if not l.is_deleted and l.category in (BPLineCategory.COSTO_FIJO, BPLineCategory.COSTO_VARIABLE)]
    total_ing = sum((l.annual_plan or 0) for l in ingreso_lines)
    total_cos = sum((l.annual_plan or 0) for l in costo_lines)
    data["computed_ingresos"] = total_ing
    data["computed_costos"] = total_cos
    data["computed_margen_pct"] = round(((total_ing - total_cos) / total_ing * 100) if total_ing > 0 else 0, 2)
    return data


@router.patch("/{bp_id}")
async def update_bp(bp_id: int, payload: BusinessPlanUpdate, db: DB, user: LeaderOrAdmin):
    result = await db.execute(
        select(BusinessPlan).where(BusinessPlan.id == bp_id, BusinessPlan.is_deleted == False)
    )
    bp = result.scalar_one_or_none()
    if not bp:
        raise HTTPException(status_code=404, detail="Business Plan no encontrado")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(bp, field, value)
    await db.flush()
    return {"id": bp.id}


@router.delete("/{bp_id}", status_code=204)
async def delete_bp(bp_id: int, db: DB, user: LeaderOrAdmin):
    result = await db.execute(
        select(BusinessPlan).where(BusinessPlan.id == bp_id, BusinessPlan.is_deleted == False)
    )
    bp = result.scalar_one_or_none()
    if not bp:
        raise HTTPException(status_code=404, detail="Business Plan no encontrado")
    bp.is_deleted = True
    await db.flush()


# ─── BPLine CRUD ──────────────────────────────────────────────────────────────

@router.get("/{bp_id}/lines")
async def list_lines(bp_id: int, db: DB, user: CurrentUser):
    result = await db.execute(
        select(BPLine)
        .where(BPLine.bp_id == bp_id, BPLine.is_deleted == False)
        .order_by(BPLine.order_index, BPLine.id)
    )
    return [_line_to_dict(l) for l in result.scalars().all()]


@router.post("/{bp_id}/lines", status_code=201)
async def create_line(bp_id: int, payload: BPLineCreate, db: DB, user: LeaderOrAdmin):
    bp_check = await db.execute(
        select(BusinessPlan).where(BusinessPlan.id == bp_id, BusinessPlan.is_deleted == False)
    )
    if not bp_check.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Business Plan no encontrado")
    data = payload.model_dump()
    if data.get("monthly_plan") and not data.get("annual_plan"):
        data["annual_plan"] = _compute_annual_plan(data["monthly_plan"])
    line = BPLine(bp_id=bp_id, **data)
    db.add(line)
    await db.flush()
    return _line_to_dict(line)


@router.patch("/{bp_id}/lines/{line_id}")
async def update_line(bp_id: int, line_id: int, payload: BPLineUpdate, db: DB, user: LeaderOrAdmin):
    result = await db.execute(
        select(BPLine).where(BPLine.id == line_id, BPLine.bp_id == bp_id, BPLine.is_deleted == False)
    )
    line = result.scalar_one_or_none()
    if not line:
        raise HTTPException(status_code=404, detail="Línea no encontrada")
    update_data = payload.model_dump(exclude_unset=True)
    if "monthly_plan" in update_data and "annual_plan" not in update_data:
        update_data["annual_plan"] = _compute_annual_plan(update_data["monthly_plan"])
    for field, value in update_data.items():
        setattr(line, field, value)
    await db.flush()
    return _line_to_dict(line)


@router.delete("/{bp_id}/lines/{line_id}", status_code=204)
async def delete_line(bp_id: int, line_id: int, db: DB, user: LeaderOrAdmin):
    result = await db.execute(
        select(BPLine).where(BPLine.id == line_id, BPLine.bp_id == bp_id, BPLine.is_deleted == False)
    )
    line = result.scalar_one_or_none()
    if not line:
        raise HTTPException(status_code=404, detail="Línea no encontrada")
    line.is_deleted = True
    await db.flush()


# ─── BPActivity CRUD ──────────────────────────────────────────────────────────

@router.get("/{bp_id}/activities")
async def list_activities(
    bp_id: int, db: DB, user: CurrentUser,
    status: Optional[str] = None,
    category: Optional[str] = None,
    priority: Optional[str] = None,
    grupo: Optional[str] = None,
):
    query = (
        select(BPActivity)
        .where(BPActivity.bp_id == bp_id, BPActivity.is_deleted == False)
        .options(
            selectinload(BPActivity.owner),
            selectinload(BPActivity.checklist),
            selectinload(BPActivity.comments),
        )
        .order_by(BPActivity.order_index, BPActivity.created_at)
    )
    if category:
        query = query.where(BPActivity.category == category)
    if priority:
        query = query.where(BPActivity.priority == priority)
    if grupo:
        query = query.where(BPActivity.grupo == grupo)
    result = await db.execute(query)
    activities = result.scalars().all()
    out = [_activity_to_dict(a) for a in activities]
    if status:
        out = [a for a in out if a["status"] == status]
    return out


@router.post("/{bp_id}/activities", status_code=201)
async def create_activity(bp_id: int, payload: BPActivityCreate, db: DB, user: LeaderOrAdmin):
    bp_check = await db.execute(
        select(BusinessPlan).where(BusinessPlan.id == bp_id, BusinessPlan.is_deleted == False)
    )
    if not bp_check.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Business Plan no encontrado")
    act = BPActivity(bp_id=bp_id, **payload.model_dump())
    db.add(act)
    await db.flush()
    result = await db.execute(
        select(BPActivity).where(BPActivity.id == act.id).options(
            selectinload(BPActivity.owner),
            selectinload(BPActivity.checklist),
            selectinload(BPActivity.comments),
        )
    )
    act = result.scalar_one()
    return _activity_to_dict(act)


@router.patch("/{bp_id}/activities/{activity_id}")
async def update_activity(bp_id: int, activity_id: int, payload: BPActivityUpdate, db: DB, user: LeaderOrAdmin):
    result = await db.execute(
        select(BPActivity)
        .where(BPActivity.id == activity_id, BPActivity.bp_id == bp_id, BPActivity.is_deleted == False)
    )
    act = result.scalar_one_or_none()
    if not act:
        raise HTTPException(status_code=404, detail="Actividad no encontrada")
    update_data = payload.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(act, field, value)
    # Auto-set VENCIDA if past due and not closed
    if act.due_date and act.due_date < date.today():
        if act.status in (BPActivityStatus.PENDIENTE, BPActivityStatus.EN_PROGRESO):
            act.status = BPActivityStatus.VENCIDA
    await db.flush()
    result2 = await db.execute(
        select(BPActivity).where(BPActivity.id == act.id).options(
            selectinload(BPActivity.owner),
            selectinload(BPActivity.checklist),
            selectinload(BPActivity.comments),
        )
    )
    act = result2.scalar_one()
    return _activity_to_dict(act)


@router.delete("/{bp_id}/activities/{activity_id}", status_code=204)
async def delete_activity(bp_id: int, activity_id: int, db: DB, user: LeaderOrAdmin):
    result = await db.execute(
        select(BPActivity)
        .where(BPActivity.id == activity_id, BPActivity.bp_id == bp_id, BPActivity.is_deleted == False)
    )
    act = result.scalar_one_or_none()
    if not act:
        raise HTTPException(status_code=404, detail="Actividad no encontrada")
    act.is_deleted = True
    await db.flush()


# ─── File Analysis (Excel or Image) ──────────────────────────────────────────

@router.post("/{bp_id}/analyze-file", status_code=201)
async def analyze_file(
    bp_id: int,
    file: UploadFile,
    db: Annotated[AsyncSession, Depends(get_db)],
    user: Annotated[User, Depends(get_current_user)],
):
    """Upload Excel OR image file, analyze with Gemini, store structured extraction."""
    bp_result = await db.execute(
        select(BusinessPlan).where(BusinessPlan.id == bp_id, BusinessPlan.is_deleted == False)
    )
    bp = bp_result.scalar_one_or_none()
    if not bp:
        raise HTTPException(status_code=404, detail="Business Plan no encontrado")

    fname = file.filename or ""
    fname_lower = fname.lower()

    # Detect file type
    is_excel = any(fname_lower.endswith(ext) for ext in EXCEL_EXTENSIONS)
    img_ext = next((ext for ext in IMAGE_MIME_TYPES if fname_lower.endswith(ext)), None)
    is_image = img_ext is not None

    if not is_excel and not is_image:
        raise HTTPException(
            status_code=400,
            detail="Formato no soportado. Use .xlsx, .xls, .xlsm, .png, .jpg, .jpeg, .gif o .webp",
        )

    contents = await file.read()
    file_size = len(contents)
    file_type = "excel" if is_excel else "image"

    parsed_data: dict = {}
    gemini_parts: list = []

    if is_excel:
        # Parse with openpyxl and build text preview for Gemini
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            sheets_data: dict = {}
            for sheet_name in wb.sheetnames[:3]:
                ws = wb[sheet_name]
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 50:
                        break
                    rows.append([str(cell) if cell is not None else "" for cell in row])
                sheets_data[sheet_name] = rows
            parsed_data = {
                "sheets": sheets_data,
                "total_sheets": len(wb.sheetnames),
                "sheet_names": list(wb.sheetnames),
            }
            wb.close()

            # Build text preview for Gemini
            preview_lines = []
            for sname, rows in list(sheets_data.items())[:3]:
                preview_lines.append(f"=== Hoja: {sname} ===")
                for row in rows[:40]:
                    line = " | ".join(r for r in row if r)
                    if line.strip():
                        preview_lines.append(line)
            preview_text = "\n".join(preview_lines[:200])
            gemini_parts = [{"text": f"Contenido del archivo Excel del BP anterior:\n\n{preview_text}"}]

        except ImportError:
            parsed_data = {"error": "openpyxl no instalado"}
            gemini_parts = [{"text": f"Archivo Excel '{fname}' (openpyxl no disponible para parsear)"}]
        except Exception as exc:
            parsed_data = {"error": str(exc)}
            gemini_parts = [{"text": f"Archivo Excel '{fname}' (error al parsear: {exc})"}]

    else:
        # Image: encode as base64 inline_data for Gemini
        mime_type = IMAGE_MIME_TYPES[img_ext]
        b64_data = base64.b64encode(contents).decode("utf-8")
        parsed_data = {"file_type": "image", "mime_type": mime_type, "size": file_size}
        gemini_parts = [
            {
                "inline_data": {
                    "mime_type": mime_type,
                    "data": b64_data,
                }
            },
            {"text": f"La imagen anterior es un dashboard/screenshot del Plan de Negocio anterior de Vanti CAS ({bp.year})."},
        ]

    # Call Gemini structured extraction
    structured_extraction = await _call_gemini_structured(db, bp.year, gemini_parts)

    # Build ai_summary from the result
    ai_summary = structured_extraction.get("summary", "")

    analysis = BPExcelAnalysis(
        bp_id=bp_id,
        filename=fname,
        file_size=file_size,
        file_type=file_type,
        parsed_data=parsed_data,
        ai_summary=ai_summary,
        ai_insights={"risks": structured_extraction.get("risks", []), "opportunities": structured_extraction.get("opportunities", [])},
        structured_extraction=structured_extraction,
        uploaded_by_id=user.id,
    )
    db.add(analysis)
    await db.flush()

    result = await db.execute(
        select(BPExcelAnalysis)
        .where(BPExcelAnalysis.id == analysis.id)
        .options(selectinload(BPExcelAnalysis.uploaded_by))
    )
    analysis = result.scalar_one()
    return _analysis_to_dict(analysis)


@router.post("/{bp_id}/apply-analysis/{analysis_id}")
async def apply_analysis(
    bp_id: int,
    analysis_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    user: Annotated[User, Depends(get_current_user)],
):
    """Apply a stored structured extraction: create BPLines, BPActivities, BPRecommendations."""
    bp_check = await db.execute(
        select(BusinessPlan).where(BusinessPlan.id == bp_id, BusinessPlan.is_deleted == False)
    )
    if not bp_check.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Business Plan no encontrado")

    analysis_result = await db.execute(
        select(BPExcelAnalysis).where(
            BPExcelAnalysis.id == analysis_id,
            BPExcelAnalysis.bp_id == bp_id,
        )
    )
    analysis = analysis_result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análisis no encontrado")

    if not analysis.structured_extraction:
        raise HTTPException(status_code=400, detail="El análisis no tiene extracción estructurada disponible")

    extraction = analysis.structured_extraction
    lines_created = 0
    activities_created = 0
    recommendations_created = 0

    # Valid category values
    valid_line_categories = {"ingreso", "costo_fijo", "costo_variable", "magnitud", "margen"}
    valid_activity_categories = {"comercial", "operativo", "financiero", "estrategico", "regulatorio", "tecnologia"}
    valid_priorities = {"critica", "alta", "media", "baja"}

    # Create financial lines
    for item in extraction.get("financial_lines", []):
        try:
            raw_cat = str(item.get("category", "magnitud")).lower()
            category = raw_cat if raw_cat in valid_line_categories else "magnitud"
            monthly_plan = item.get("monthly_plan") or {}
            # Ensure keys are strings and values are numbers
            clean_monthly = {}
            for k, v in monthly_plan.items():
                try:
                    clean_monthly[str(k)] = float(v) if v is not None else 0.0
                except (TypeError, ValueError):
                    clean_monthly[str(k)] = 0.0
            annual_plan = _compute_annual_plan(clean_monthly)
            line = BPLine(
                bp_id=bp_id,
                category=category,
                subcategory=item.get("subcategory"),
                name=str(item.get("name", "Línea sin nombre"))[:200],
                unit=str(item.get("unit", "COP"))[:30],
                monthly_plan=clean_monthly if clean_monthly else None,
                annual_plan=annual_plan,
                is_ai_generated=True,
                ai_confidence=int(item.get("ai_confidence", 70)) if item.get("ai_confidence") is not None else 70,
                ai_rationale=str(item.get("ai_rationale", ""))[:500] if item.get("ai_rationale") else None,
            )
            db.add(line)
            lines_created += 1
        except Exception:
            continue

    # Create custom_metrics as magnitud lines if present
    for item in extraction.get("custom_metrics", []):
        try:
            monthly_plan = item.get("monthly_plan") or {}
            clean_monthly = {}
            for k, v in monthly_plan.items():
                try:
                    clean_monthly[str(k)] = float(v) if v is not None else 0.0
                except (TypeError, ValueError):
                    clean_monthly[str(k)] = 0.0
            line = BPLine(
                bp_id=bp_id,
                category="magnitud",
                subcategory=item.get("subcategory"),
                name=str(item.get("name", "Métrica"))[:200],
                unit=str(item.get("unit", "unidades"))[:30],
                monthly_plan=clean_monthly if clean_monthly else None,
                annual_plan=_compute_annual_plan(clean_monthly),
                is_ai_generated=True,
                ai_confidence=int(item.get("ai_confidence", 70)) if item.get("ai_confidence") is not None else 70,
                ai_rationale=str(item.get("ai_rationale", ""))[:500] if item.get("ai_rationale") else None,
                line_metadata={"source": "custom_metric"},
            )
            db.add(line)
            lines_created += 1
        except Exception:
            continue

    # Create activities
    for item in extraction.get("activities", []):
        try:
            raw_cat = str(item.get("category", "operativo")).lower()
            category = raw_cat if raw_cat in valid_activity_categories else "operativo"
            raw_prio = str(item.get("priority", "media")).lower()
            priority = raw_prio if raw_prio in valid_priorities else "media"

            due_date = None
            raw_due = item.get("due_date")
            if raw_due and raw_due != "null":
                try:
                    from datetime import date as _date
                    due_date = _date.fromisoformat(str(raw_due))
                except (ValueError, TypeError):
                    due_date = None

            act = BPActivity(
                bp_id=bp_id,
                title=str(item.get("title", "Actividad sin título"))[:300],
                description=str(item.get("description", ""))[:1000] if item.get("description") else None,
                category=category,
                priority=priority,
                status="pendiente",
                due_date=due_date,
                notes=str(item.get("notes", ""))[:500] if item.get("notes") else None,
            )
            db.add(act)
            activities_created += 1
        except Exception:
            continue

    # Create recommendations
    valid_rec_categories = {"comercial", "financiero", "operativo", "estrategico", "riesgo", "oportunidad"}
    valid_impact_levels = {"alto", "medio", "bajo"}

    for item in extraction.get("recommendations", []):
        try:
            raw_cat = str(item.get("category", "estrategico")).lower()
            category = raw_cat if raw_cat in valid_rec_categories else "estrategico"
            raw_prio = str(item.get("priority", "media")).lower()
            priority = raw_prio if raw_prio in valid_priorities else "media"
            raw_impact = str(item.get("impact_level", "medio")).lower() if item.get("impact_level") else None
            impact_level = raw_impact if raw_impact in valid_impact_levels else "medio"

            rec = BPRecommendation(
                bp_id=bp_id,
                source="ai",
                category=category,
                title=str(item.get("title", "Recomendación"))[:300],
                description=str(item.get("description", ""))[:2000] if item.get("description") else None,
                priority=priority,
                status="pendiente",
                impact_level=impact_level,
                is_ai_generated=True,
            )
            db.add(rec)
            recommendations_created += 1
        except Exception:
            continue

    # Mark analysis as applied
    analysis.applied_at = datetime.utcnow()
    await db.flush()

    return {
        "lines_created": lines_created,
        "activities_created": activities_created,
        "recommendations_created": recommendations_created,
    }


# ─── Legacy Excel Analysis (backward compat) ─────────────────────────────────

@router.post("/{bp_id}/analyze-excel", status_code=201)
async def analyze_excel(
    bp_id: int,
    file: UploadFile,
    db: Annotated[AsyncSession, Depends(get_db)],
    user: Annotated[User, Depends(get_current_user)],
):
    """Legacy endpoint — delegates to analyze_file for backward compatibility."""
    return await analyze_file(bp_id, file, db, user)


@router.get("/{bp_id}/analyses")
async def list_analyses(bp_id: int, db: DB, user: CurrentUser):
    result = await db.execute(
        select(BPExcelAnalysis)
        .where(BPExcelAnalysis.bp_id == bp_id)
        .options(selectinload(BPExcelAnalysis.uploaded_by))
        .order_by(BPExcelAnalysis.uploaded_at.desc())
    )
    return [_analysis_to_dict(a) for a in result.scalars().all()]


# ─── Recommendations CRUD ─────────────────────────────────────────────────────

@router.get("/{bp_id}/recommendations")
async def list_recommendations(
    bp_id: int,
    db: DB,
    user: CurrentUser,
    status: Optional[str] = None,
    category: Optional[str] = None,
):
    query = (
        select(BPRecommendation)
        .where(BPRecommendation.bp_id == bp_id, BPRecommendation.is_deleted == False)
        .order_by(BPRecommendation.created_at.desc())
    )
    if status:
        query = query.where(BPRecommendation.status == status)
    if category:
        query = query.where(BPRecommendation.category == category)
    result = await db.execute(query)
    return [_recommendation_to_dict(r) for r in result.scalars().all()]


@router.post("/{bp_id}/recommendations", status_code=201)
async def create_recommendation(bp_id: int, payload: BPRecommendationCreate, db: DB, user: LeaderOrAdmin):
    bp_check = await db.execute(
        select(BusinessPlan).where(BusinessPlan.id == bp_id, BusinessPlan.is_deleted == False)
    )
    if not bp_check.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Business Plan no encontrado")

    rec = BPRecommendation(
        bp_id=bp_id,
        source=payload.source,
        category=payload.category,
        title=payload.title,
        description=payload.description,
        priority=payload.priority,
        impact_level=payload.impact_level,
        is_ai_generated=(payload.source == "ai"),
        status="pendiente",
    )
    db.add(rec)
    await db.flush()
    return _recommendation_to_dict(rec)


@router.patch("/{bp_id}/recommendations/{rec_id}")
async def update_recommendation(
    bp_id: int, rec_id: int, payload: BPRecommendationUpdate, db: DB, user: LeaderOrAdmin
):
    result = await db.execute(
        select(BPRecommendation).where(
            BPRecommendation.id == rec_id,
            BPRecommendation.bp_id == bp_id,
            BPRecommendation.is_deleted == False,
        )
    )
    rec = result.scalar_one_or_none()
    if not rec:
        raise HTTPException(status_code=404, detail="Recomendación no encontrada")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(rec, field, value)
    await db.flush()
    return _recommendation_to_dict(rec)


@router.delete("/{bp_id}/recommendations/{rec_id}", status_code=204)
async def delete_recommendation(bp_id: int, rec_id: int, db: DB, user: LeaderOrAdmin):
    result = await db.execute(
        select(BPRecommendation).where(
            BPRecommendation.id == rec_id,
            BPRecommendation.bp_id == bp_id,
            BPRecommendation.is_deleted == False,
        )
    )
    rec = result.scalar_one_or_none()
    if not rec:
        raise HTTPException(status_code=404, detail="Recomendación no encontrada")
    rec.is_deleted = True
    await db.flush()


# ─── Checklist CRUD ───────────────────────────────────────────────────────────

async def _get_activity_or_404(db: AsyncSession, bp_id: int, act_id: int) -> BPActivity:
    result = await db.execute(
        select(BPActivity).where(
            BPActivity.id == act_id,
            BPActivity.bp_id == bp_id,
            BPActivity.is_deleted == False,
        )
    )
    act = result.scalar_one_or_none()
    if not act:
        raise HTTPException(status_code=404, detail="Actividad no encontrada")
    return act


async def _recalculate_activity_progress(db: AsyncSession, act_id: int) -> None:
    """Recalculate activity progress from checklist completion %."""
    result = await db.execute(
        select(BPChecklist).where(BPChecklist.activity_id == act_id)
    )
    items = result.scalars().all()
    if not items:
        return
    done = sum(1 for i in items if i.is_completed)
    total = len(items)
    progress = round(done / total * 100) if total > 0 else 0
    act_result = await db.execute(
        select(BPActivity).where(BPActivity.id == act_id)
    )
    act = act_result.scalar_one_or_none()
    if act:
        act.progress = progress


@router.get("/{bp_id}/activities/{act_id}/checklist")
async def list_checklist(bp_id: int, act_id: int, db: DB, user: CurrentUser):
    await _get_activity_or_404(db, bp_id, act_id)
    result = await db.execute(
        select(BPChecklist)
        .where(BPChecklist.activity_id == act_id)
        .order_by(BPChecklist.order_index, BPChecklist.id)
    )
    items = result.scalars().all()
    return [
        {
            "id": i.id,
            "activity_id": i.activity_id,
            "title": i.title,
            "is_completed": i.is_completed,
            "completed_at": i.completed_at,
            "completed_by_id": i.completed_by_id,
            "order_index": i.order_index,
            "created_at": i.created_at,
        }
        for i in items
    ]


@router.post("/{bp_id}/activities/{act_id}/checklist", status_code=201)
async def add_checklist_item(
    bp_id: int, act_id: int, payload: BPChecklistItemCreate, db: DB, user: LeaderOrAdmin
):
    await _get_activity_or_404(db, bp_id, act_id)
    item = BPChecklist(
        activity_id=act_id,
        title=payload.title,
        order_index=payload.order_index,
    )
    db.add(item)
    await db.flush()
    return {
        "id": item.id,
        "activity_id": item.activity_id,
        "title": item.title,
        "is_completed": item.is_completed,
        "completed_at": item.completed_at,
        "completed_by_id": item.completed_by_id,
        "order_index": item.order_index,
        "created_at": item.created_at,
    }


@router.patch("/{bp_id}/activities/{act_id}/checklist/{item_id}")
async def update_checklist_item(
    bp_id: int, act_id: int, item_id: int, payload: BPChecklistItemUpdate,
    db: DB, user: LeaderOrAdmin
):
    await _get_activity_or_404(db, bp_id, act_id)
    result = await db.execute(
        select(BPChecklist).where(
            BPChecklist.id == item_id,
            BPChecklist.activity_id == act_id,
        )
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Ítem de checklist no encontrado")

    update_data = payload.model_dump(exclude_unset=True)
    if "is_completed" in update_data:
        if update_data["is_completed"] and not item.is_completed:
            item.completed_at = datetime.utcnow()
            item.completed_by_id = user.id
        elif not update_data["is_completed"]:
            item.completed_at = None
            item.completed_by_id = None

    for field, value in update_data.items():
        setattr(item, field, value)

    await db.flush()
    await _recalculate_activity_progress(db, act_id)
    await db.flush()

    return {
        "id": item.id,
        "activity_id": item.activity_id,
        "title": item.title,
        "is_completed": item.is_completed,
        "completed_at": item.completed_at,
        "completed_by_id": item.completed_by_id,
        "order_index": item.order_index,
        "created_at": item.created_at,
    }


@router.delete("/{bp_id}/activities/{act_id}/checklist/{item_id}", status_code=204)
async def delete_checklist_item(
    bp_id: int, act_id: int, item_id: int, db: DB, user: LeaderOrAdmin
):
    await _get_activity_or_404(db, bp_id, act_id)
    result = await db.execute(
        select(BPChecklist).where(
            BPChecklist.id == item_id,
            BPChecklist.activity_id == act_id,
        )
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Ítem de checklist no encontrado")
    await db.delete(item)
    await db.flush()
    await _recalculate_activity_progress(db, act_id)
    await db.flush()


# ─── Comments CRUD ────────────────────────────────────────────────────────────

@router.get("/{bp_id}/activities/{act_id}/comments")
async def list_comments(bp_id: int, act_id: int, db: DB, user: CurrentUser):
    await _get_activity_or_404(db, bp_id, act_id)
    result = await db.execute(
        select(BPComment)
        .where(BPComment.activity_id == act_id, BPComment.is_deleted == False)
        .options(selectinload(BPComment.author))
        .order_by(BPComment.created_at)
    )
    comments = result.scalars().all()
    return [
        {
            "id": c.id,
            "activity_id": c.activity_id,
            "author_id": c.author_id,
            "author_name": c.author.full_name if c.author else "",
            "content": c.content,
            "created_at": c.created_at,
            "updated_at": c.updated_at,
        }
        for c in comments
    ]


@router.post("/{bp_id}/activities/{act_id}/comments", status_code=201)
async def add_comment(
    bp_id: int, act_id: int, payload: BPCommentCreate, db: DB, user: CurrentUser
):
    await _get_activity_or_404(db, bp_id, act_id)
    comment = BPComment(
        activity_id=act_id,
        author_id=user.id,
        content=payload.content,
    )
    db.add(comment)
    await db.flush()
    result = await db.execute(
        select(BPComment).where(BPComment.id == comment.id).options(selectinload(BPComment.author))
    )
    comment = result.scalar_one()
    return {
        "id": comment.id,
        "activity_id": comment.activity_id,
        "author_id": comment.author_id,
        "author_name": comment.author.full_name if comment.author else "",
        "content": comment.content,
        "created_at": comment.created_at,
        "updated_at": comment.updated_at,
    }


@router.delete("/{bp_id}/activities/{act_id}/comments/{comment_id}", status_code=204)
async def delete_comment(
    bp_id: int, act_id: int, comment_id: int, db: DB, user: CurrentUser
):
    await _get_activity_or_404(db, bp_id, act_id)
    result = await db.execute(
        select(BPComment).where(
            BPComment.id == comment_id,
            BPComment.activity_id == act_id,
            BPComment.is_deleted == False,
        )
    )
    comment = result.scalar_one_or_none()
    if not comment:
        raise HTTPException(status_code=404, detail="Comentario no encontrado")
    # Only author or admin can delete
    is_admin = getattr(user, "role", None) in ("admin", "leader")
    if comment.author_id != user.id and not is_admin:
        raise HTTPException(status_code=403, detail="No tienes permiso para eliminar este comentario")
    comment.is_deleted = True
    await db.flush()


# ─── Milestones CRUD ──────────────────────────────────────────────────────────

@router.get("/{bp_id}/milestones")
async def list_milestones(bp_id: int, db: DB, user: CurrentUser):
    result = await db.execute(
        select(BPMilestone)
        .where(BPMilestone.bp_id == bp_id, BPMilestone.is_deleted == False)
        .order_by(BPMilestone.target_date, BPMilestone.order_index)
    )
    milestones = result.scalars().all()
    return [
        {
            "id": m.id,
            "bp_id": m.bp_id,
            "title": m.title,
            "description": m.description,
            "target_date": str(m.target_date),
            "status": m.status,
            "color": m.color,
            "order_index": m.order_index,
            "created_at": m.created_at,
        }
        for m in milestones
    ]


@router.post("/{bp_id}/milestones", status_code=201)
async def create_milestone(bp_id: int, payload: BPMilestoneCreate, db: DB, user: LeaderOrAdmin):
    bp_check = await db.execute(
        select(BusinessPlan).where(BusinessPlan.id == bp_id, BusinessPlan.is_deleted == False)
    )
    if not bp_check.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Business Plan no encontrado")
    ms = BPMilestone(bp_id=bp_id, **payload.model_dump())
    db.add(ms)
    await db.flush()
    return {
        "id": ms.id,
        "bp_id": ms.bp_id,
        "title": ms.title,
        "description": ms.description,
        "target_date": str(ms.target_date),
        "status": ms.status,
        "color": ms.color,
        "order_index": ms.order_index,
        "created_at": ms.created_at,
    }


@router.patch("/{bp_id}/milestones/{ms_id}")
async def update_milestone(
    bp_id: int, ms_id: int, payload: BPMilestoneUpdate, db: DB, user: LeaderOrAdmin
):
    result = await db.execute(
        select(BPMilestone).where(
            BPMilestone.id == ms_id,
            BPMilestone.bp_id == bp_id,
            BPMilestone.is_deleted == False,
        )
    )
    ms = result.scalar_one_or_none()
    if not ms:
        raise HTTPException(status_code=404, detail="Hito no encontrado")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(ms, field, value)
    await db.flush()
    return {
        "id": ms.id,
        "bp_id": ms.bp_id,
        "title": ms.title,
        "description": ms.description,
        "target_date": str(ms.target_date),
        "status": ms.status,
        "color": ms.color,
        "order_index": ms.order_index,
        "created_at": ms.created_at,
    }


@router.delete("/{bp_id}/milestones/{ms_id}", status_code=204)
async def delete_milestone(bp_id: int, ms_id: int, db: DB, user: LeaderOrAdmin):
    result = await db.execute(
        select(BPMilestone).where(
            BPMilestone.id == ms_id,
            BPMilestone.bp_id == bp_id,
            BPMilestone.is_deleted == False,
        )
    )
    ms = result.scalar_one_or_none()
    if not ms:
        raise HTTPException(status_code=404, detail="Hito no encontrado")
    ms.is_deleted = True
    await db.flush()


# ─── Timeline (Gantt) ─────────────────────────────────────────────────────────

@router.get("/{bp_id}/timeline")
async def get_timeline(bp_id: int, db: DB, user: CurrentUser):
    """Return structured Gantt data for the BP timeline view."""
    bp_result = await db.execute(
        select(BusinessPlan).where(BusinessPlan.id == bp_id, BusinessPlan.is_deleted == False)
    )
    bp = bp_result.scalar_one_or_none()
    if not bp:
        raise HTTPException(status_code=404, detail="Business Plan no encontrado")

    # Load activities with related data
    acts_result = await db.execute(
        select(BPActivity)
        .where(BPActivity.bp_id == bp_id, BPActivity.is_deleted == False)
        .options(
            selectinload(BPActivity.owner),
            selectinload(BPActivity.checklist),
            selectinload(BPActivity.comments),
        )
        .order_by(BPActivity.order_index, BPActivity.created_at)
    )
    activities = acts_result.scalars().all()

    # Load milestones
    ms_result = await db.execute(
        select(BPMilestone)
        .where(BPMilestone.bp_id == bp_id, BPMilestone.is_deleted == False)
        .order_by(BPMilestone.target_date)
    )
    milestones = ms_result.scalars().all()

    today = date.today()
    all_dates = []

    act_dicts = []
    by_owner: dict = {}
    total = len(activities)
    completed = 0
    in_progress = 0
    overdue = 0
    on_track = 0

    for act in activities:
        start = act.start_date or act.created_at.date()
        end = act.due_date  # may be None
        owner_name = act.owner.full_name if act.owner else "Sin asignar"
        checklist = act.checklist or []
        checklist_total = len(checklist)
        checklist_done = sum(1 for i in checklist if i.is_completed)

        displayed_status = act.status.value if hasattr(act.status, "value") else act.status
        is_overdue_act = _is_overdue(act)
        if is_overdue_act and act.status == BPActivityStatus.PENDIENTE:
            displayed_status = BPActivityStatus.VENCIDA.value

        act_dicts.append({
            "id": act.id,
            "title": act.title,
            "category": act.category.value if hasattr(act.category, "value") else act.category,
            "priority": act.priority.value if hasattr(act.priority, "value") else act.priority,
            "status": displayed_status,
            "owner_id": act.owner_id,
            "owner_name": owner_name,
            "start_date": str(start),
            "due_date": str(act.due_date) if act.due_date else None,
            "progress": act.progress,
            "is_milestone": act.is_milestone,
            "depends_on_id": act.depends_on_id,
            "checklist_total": checklist_total,
            "checklist_done": checklist_done,
            "tags": act.tags,
            "is_overdue": is_overdue_act,
            "estimated_hours": act.estimated_hours,
            "actual_hours": act.actual_hours,
        })

        all_dates.append(start)
        if act.due_date:
            all_dates.append(act.due_date)

        if act.status == BPActivityStatus.COMPLETADA:
            completed += 1
        elif act.status == BPActivityStatus.EN_PROGRESO:
            in_progress += 1
        if is_overdue_act:
            overdue += 1
        elif act.status not in (BPActivityStatus.COMPLETADA, BPActivityStatus.CANCELADA):
            on_track += 1

        # Group by owner
        by_owner.setdefault(owner_name, []).append(act.id)

    ms_dicts = []
    for m in milestones:
        all_dates.append(m.target_date)
        ms_dicts.append({
            "id": m.id,
            "title": m.title,
            "description": m.description,
            "target_date": str(m.target_date),
            "status": m.status,
            "color": m.color,
            "order_index": m.order_index,
        })

    date_range = {
        "min": str(min(all_dates)) if all_dates else str(today),
        "max": str(max(all_dates)) if all_dates else str(today),
    }

    return {
        "activities": act_dicts,
        "milestones": ms_dicts,
        "date_range": date_range,
        "by_owner": by_owner,
        "stats": {
            "total": total,
            "completed": completed,
            "in_progress": in_progress,
            "overdue": overdue,
            "on_track": on_track,
        },
    }


# ─── Reminder Check ───────────────────────────────────────────────────────────

@router.post("/{bp_id}/check-reminders")
async def check_reminders(bp_id: int, db: DB, user: CurrentUser):
    """Check for activities near due date or overdue, send notifications."""
    bp_result = await db.execute(
        select(BusinessPlan).where(BusinessPlan.id == bp_id, BusinessPlan.is_deleted == False)
    )
    bp = bp_result.scalar_one_or_none()
    if not bp:
        raise HTTPException(status_code=404, detail="Business Plan no encontrado")

    acts_result = await db.execute(
        select(BPActivity).where(
            BPActivity.bp_id == bp_id,
            BPActivity.is_deleted == False,
            BPActivity.status.notin_([BPActivityStatus.COMPLETADA, BPActivityStatus.CANCELADA]),
        )
    )
    activities = acts_result.scalars().all()

    today = date.today()
    now = datetime.utcnow()
    reminders_sent = 0
    overdue_notified = 0

    for act in activities:
        if not act.owner_id:
            continue

        # Check overdue
        if act.due_date and act.due_date < today:
            # Send overdue notification if not sent recently (within 7 days)
            should_send = (
                act.reminder_sent_at is None
                or (now - act.reminder_sent_at).days >= 7
            )
            if should_send:
                notif = Notification(
                    user_id=act.owner_id,
                    title=f"Tarea BP vencida: {act.title}",
                    message=(
                        f"La actividad '{act.title}' del BP {bp.year} venció el "
                        f"{act.due_date.strftime('%d/%m/%Y')}. Por favor actualiza su estado."
                    ),
                    notification_type="error",
                    entity_type="bp_activity",
                    entity_id=act.id,
                )
                db.add(notif)
                act.reminder_sent_at = now
                overdue_notified += 1
            continue

        # Check upcoming (within reminder_days_before)
        if act.due_date:
            days_until = (act.due_date - today).days
            if 0 <= days_until <= act.reminder_days_before:
                should_send = (
                    act.reminder_sent_at is None
                    or (now - act.reminder_sent_at).days >= 7
                )
                if should_send:
                    notif = Notification(
                        user_id=act.owner_id,
                        title=f"Tarea BP próxima a vencer: {act.title}",
                        message=(
                            f"La actividad '{act.title}' del BP {bp.year} vence el "
                            f"{act.due_date.strftime('%d/%m/%Y')} ({days_until} día(s) restante(s))."
                        ),
                        notification_type="warning",
                        entity_type="bp_activity",
                        entity_id=act.id,
                    )
                    db.add(notif)
                    act.reminder_sent_at = now
                    reminders_sent += 1

    await db.flush()
    return {"reminders_sent": reminders_sent, "overdue_notified": overdue_notified}
